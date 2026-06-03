from __future__ import annotations

import tempfile
import unittest
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.portal_workbook import load_portal_issue_rows, sha256_file, summarize_portal_issue_rows


def build_portal_workbook(path: Path, amount_row_1: str | None = "247.52", amount_row_2: str | None = "182.18") -> None:
    workbook = Workbook()
    basic = workbook.active
    basic.title = "1-发票基本信息"
    detail = workbook.create_sheet("2-发票明细信息")

    basic_headers = {
        "发票流水号": 1,
        "购买方名称": 5,
        "购买方纳税人识别号": 6,
        "购买方邮箱": 7,
    }
    detail_headers = {
        "发票流水号": 1,
        "项目名称": 2,
        "金额": 4,
        "税率": 5,
    }
    for header, column in basic_headers.items():
        basic.cell(3, column).value = header
    for header, column in detail_headers.items():
        detail.cell(3, column).value = header

    basic_rows = [
        ("1", "洪玮颖", "", "1611461905@qq.com"),
        ("2", "陈丹洁", "", "645368635@qq.com"),
    ]
    detail_rows = [
        ("1", "*餐饮服务*餐费", amount_row_2, "0.01"),
        ("2", "*餐饮服务*餐费", amount_row_1, "0.01"),
    ]
    for index, row in enumerate(basic_rows, start=4):
        basic.cell(index, basic_headers["发票流水号"]).value = row[0]
        basic.cell(index, basic_headers["购买方名称"]).value = row[1]
        basic.cell(index, basic_headers["购买方纳税人识别号"]).value = row[2]
        basic.cell(index, basic_headers["购买方邮箱"]).value = row[3]
    for index, row in enumerate(detail_rows, start=4):
        detail.cell(index, detail_headers["发票流水号"]).value = row[0]
        detail.cell(index, detail_headers["项目名称"]).value = row[1]
        detail.cell(index, detail_headers["金额"]).value = row[2]
        detail.cell(index, detail_headers["税率"]).value = row[3]

    workbook.save(path)


class PortalWorkbookTests(unittest.TestCase):
    def test_load_portal_issue_rows_parses_amounts_and_tax(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "portal.xlsx"
            build_portal_workbook(workbook_path)

            rows = load_portal_issue_rows(workbook_path)
            summary = summarize_portal_issue_rows(rows)

            self.assertEqual(2, len(rows))
            self.assertEqual("洪玮颖", rows[0].buyer_name)
            self.assertEqual(Decimal("182.18"), rows[0].amount_excluding_tax)
            self.assertEqual(Decimal("1.82"), rows[0].tax_amount)
            self.assertEqual(Decimal("184.00"), rows[0].amount_including_tax)
            self.assertEqual("陈丹洁", rows[1].buyer_name)
            self.assertEqual(Decimal("434.00"), summary.total_amount_including_tax)
            self.assertEqual(64, len(sha256_file(workbook_path)))

    def test_load_portal_issue_rows_blocks_on_empty_amount(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "portal.xlsx"
            build_portal_workbook(workbook_path, amount_row_2=None)

            with self.assertRaisesRegex(ValueError, "missing 金额"):
                load_portal_issue_rows(workbook_path, block_on_empty_amount=True)

    def test_load_portal_issue_rows_returns_empty_list_for_cleared_template(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "portal.xlsx"
            build_portal_workbook(workbook_path)

            workbook = load_workbook(workbook_path)
            for sheet_name in ("1-发票基本信息", "2-发票明细信息"):
                sheet = workbook[sheet_name]
                for row in range(4, sheet.max_row + 1):
                    for column in range(1, sheet.max_column + 1):
                        sheet.cell(row, column).value = None
            workbook.save(workbook_path)

            rows = load_portal_issue_rows(workbook_path, block_on_empty_amount=False)
            summary = summarize_portal_issue_rows(rows)

            self.assertEqual([], rows)
            self.assertEqual(0, summary.row_count)
            self.assertEqual(Decimal("0.00"), summary.total_amount_including_tax)
