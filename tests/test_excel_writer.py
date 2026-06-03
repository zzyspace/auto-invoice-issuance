from __future__ import annotations

import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.excel_writer import InvoiceExcelWriter
from app.models import NormalizedInvoice, StoreConfig


class ExcelWriterTests(unittest.TestCase):
    def test_writer_creates_backup_and_rewrites_incremental_rows(self) -> None:
        repo_root = Path(__file__).resolve().parents[1]
        template_path = repo_root / "(V260401版)批量开票-导入开票模板.xlsx"
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            output_path = tmp_path / "output" / "store_a.xlsx"
            output_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(template_path, output_path)

            writer = InvoiceExcelWriter(template_path=template_path, backups_root=tmp_path / "backups")
            store = StoreConfig(
                store_key="store_a",
                store_name="门店A",
                survey_id="22512014",
                output_xlsx_path=output_path,
                initial_last_processed_id=307,
                enabled=True,
            )
            invoices = [
                NormalizedInvoice(
                    source_submission_id=308,
                    invoice_serial="1",
                    invoice_title="吴翔",
                    is_natural_person=True,
                    tax_id=None,
                    email="a@example.com",
                    amount_text="246",
                    remark="",
                ),
                NormalizedInvoice(
                    source_submission_id=309,
                    invoice_serial="2",
                    invoice_title="深圳易思商务咨询有限公司厦门分公司",
                    is_natural_person=False,
                    tax_id="AB12345",
                    email="b@example.com",
                    amount_text="318.5",
                    remark="",
                ),
            ]

            result = writer.write_store_workbook(store, invoices)

            self.assertIsNotNone(result.backup_path)
            self.assertTrue(result.backup_path and result.backup_path.exists())
            workbook = load_workbook(output_path)
            basic_sheet = workbook["1-发票基本信息"]
            detail_sheet = workbook["2-发票明细信息"]

            self.assertEqual("1", basic_sheet["A4"].value)
            self.assertEqual("吴翔", basic_sheet["F4"].value)
            self.assertEqual("是", basic_sheet["E4"].value)
            self.assertEqual("2", basic_sheet["A5"].value)
            self.assertEqual("否", basic_sheet["E5"].value)
            self.assertEqual("AB12345", basic_sheet["G5"].value)
            self.assertIsNone(basic_sheet["A6"].value)

            self.assertEqual("1", detail_sheet["A4"].value)
            self.assertEqual("餐费", detail_sheet["B4"].value)
            self.assertEqual("246", detail_sheet["H4"].value)
            self.assertEqual("318.5", detail_sheet["H5"].value)
            self.assertIsNone(detail_sheet["A6"].value)

    def test_writer_clears_existing_rows_when_invoices_are_empty(self) -> None:
        repo_root = Path(__file__).resolve().parents[1]
        template_path = repo_root / "(V260401版)批量开票-导入开票模板.xlsx"
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            output_path = tmp_path / "output" / "store_a.xlsx"
            output_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(template_path, output_path)

            writer = InvoiceExcelWriter(template_path=template_path, backups_root=tmp_path / "backups")
            store = StoreConfig(
                store_key="store_a",
                store_name="门店A",
                survey_id="22512014",
                output_xlsx_path=output_path,
                initial_last_processed_id=307,
                enabled=True,
            )
            seeded_workbook = load_workbook(output_path)
            seeded_basic_sheet = seeded_workbook["1-发票基本信息"]
            seeded_detail_sheet = seeded_workbook["2-发票明细信息"]
            seeded_basic_sheet["A4"] = "existing"
            seeded_basic_sheet["F4"] = "旧数据"
            seeded_detail_sheet["A4"] = "existing"
            seeded_detail_sheet["H4"] = "100"
            seeded_workbook.save(output_path)

            result = writer.write_store_workbook(store, [])

            self.assertIsNotNone(result.backup_path)
            self.assertTrue(result.backup_path and result.backup_path.exists())
            workbook = load_workbook(output_path)
            basic_sheet = workbook["1-发票基本信息"]
            detail_sheet = workbook["2-发票明细信息"]

            self.assertIsNone(basic_sheet["A4"].value)
            self.assertIsNone(basic_sheet["F4"].value)
            self.assertIsNone(detail_sheet["A4"].value)
            self.assertIsNone(detail_sheet["H4"].value)
