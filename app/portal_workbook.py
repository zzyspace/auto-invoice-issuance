from __future__ import annotations

import hashlib
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook

from app.models import PortalIssueRow

BASIC_SHEET = "1-发票基本信息"
DETAIL_SHEET = "2-发票明细信息"
HEADER_ROW = 3
DATA_START_ROW = 4


@dataclass(frozen=True)
class PortalIssueWorkbookSummary:
    row_count: int
    total_amount_excluding_tax: Decimal
    total_tax_amount: Decimal
    total_amount_including_tax: Decimal


def load_portal_issue_rows(path: Path, block_on_empty_amount: bool = True) -> list[PortalIssueRow]:
    workbook = load_workbook(path, data_only=True)
    basic_sheet = workbook[BASIC_SHEET]
    detail_sheet = workbook[DETAIL_SHEET]

    basic_headers = _header_map(basic_sheet)
    detail_headers = _header_map(detail_sheet)

    basic_rows = {
        row["发票流水号"]: row
        for row in _sheet_rows(
            basic_sheet,
            basic_headers,
            ("发票流水号", "购买方名称", "购买方纳税人识别号", "购买方邮箱"),
        )
    }
    detail_rows = _sheet_rows(
        detail_sheet,
        detail_headers,
        ("发票流水号", "金额", "税率"),
    )

    issue_rows: list[PortalIssueRow] = []
    for detail_row in detail_rows:
        invoice_serial = detail_row["发票流水号"]
        basic_row = basic_rows.get(invoice_serial)
        if basic_row is None:
            raise ValueError(f"Workbook row {invoice_serial} is missing from {BASIC_SHEET}.")
        amount_text = detail_row["金额"]
        if amount_text is None or str(amount_text).strip() == "":
            if block_on_empty_amount:
                raise ValueError(f"Workbook row {invoice_serial} is missing 金额 and cannot be submitted.")
            continue

        amount_excluding_tax = _parse_decimal(amount_text, label=f"row {invoice_serial} 金额")
        if amount_excluding_tax <= Decimal("0"):
            raise ValueError(f"Workbook row {invoice_serial} has non-positive 金额: {amount_excluding_tax}")
        tax_rate = _parse_decimal(detail_row["税率"] or "0", label=f"row {invoice_serial} 税率")
        tax_amount = (amount_excluding_tax * tax_rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        amount_including_tax = (amount_excluding_tax + tax_amount).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
        issue_rows.append(
            PortalIssueRow(
                invoice_serial=invoice_serial,
                buyer_name=str(basic_row["购买方名称"] or "").strip(),
                buyer_tax_id=_optional_text(basic_row["购买方纳税人识别号"]),
                buyer_email=_optional_text(basic_row["购买方邮箱"]),
                amount_excluding_tax=amount_excluding_tax,
                tax_rate=tax_rate,
                tax_amount=tax_amount,
                amount_including_tax=amount_including_tax,
            )
        )

    issue_rows.sort(key=lambda row: _sort_key(row.invoice_serial))
    return issue_rows


def summarize_portal_issue_rows(rows: list[PortalIssueRow]) -> PortalIssueWorkbookSummary:
    total_amount_excluding_tax = sum((row.amount_excluding_tax for row in rows), Decimal("0"))
    total_tax_amount = sum((row.tax_amount for row in rows), Decimal("0"))
    total_amount_including_tax = sum((row.amount_including_tax for row in rows), Decimal("0"))
    return PortalIssueWorkbookSummary(
        row_count=len(rows),
        total_amount_excluding_tax=total_amount_excluding_tax.quantize(Decimal("0.01")),
        total_tax_amount=total_tax_amount.quantize(Decimal("0.01")),
        total_amount_including_tax=total_amount_including_tax.quantize(Decimal("0.01")),
    )


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _header_map(sheet: object) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        value = sheet.cell(HEADER_ROW, column).value
        if value is None:
            continue
        headers[str(value).strip()] = column
    return headers


def _sheet_rows(sheet: object, header_map: dict[str, int], required_headers: tuple[str, ...]) -> list[dict[str, str | None]]:
    rows: list[dict[str, str | None]] = []
    for header in required_headers:
        if header not in header_map:
            raise ValueError(f"Workbook sheet '{sheet.title}' is missing header: {header}")
    for row_index in range(DATA_START_ROW, sheet.max_row + 1):
        values = {
            header: sheet.cell(row_index, header_map[header]).value
            for header in required_headers
        }
        if all(value in (None, "") for value in values.values()):
            continue
        normalized = {
            header: str(value).strip() if value is not None and str(value).strip() != "" else None
            for header, value in values.items()
        }
        if not normalized[required_headers[0]]:
            raise ValueError(f"Workbook sheet '{sheet.title}' row {row_index} is missing {required_headers[0]}")
        rows.append(normalized)
    return rows


def _parse_decimal(value: str | None, label: str) -> Decimal:
    if value is None:
        raise ValueError(f"{label} is missing")
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{label} is not a valid decimal: {value!r}") from exc


def _optional_text(value: str | None) -> str | None:
    if value is None:
        return None
    stripped = str(value).strip()
    return stripped or None


def _sort_key(invoice_serial: str) -> tuple[int, str]:
    if invoice_serial.isdigit():
        return (0, f"{int(invoice_serial):020d}")
    return (1, invoice_serial)
