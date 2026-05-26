from __future__ import annotations

import shutil
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook

from app.models import NormalizedInvoice, StoreConfig
from app.utils import BackupResult, clone_row_style, ensure_parent_dir

DATA_START_ROW = 4
HEADER_ROW = 3
TEMPLATE_ROW = 4


class InvoiceExcelWriter:
    def __init__(self, template_path: Path, backups_root: Path) -> None:
        self.template_path = template_path
        self.backups_root = backups_root

    def write_store_workbook(
        self,
        store: StoreConfig,
        invoices: list[NormalizedInvoice],
    ) -> BackupResult:
        backup_result = self._prepare_output(store)
        workbook = load_workbook(self.template_path)
        basic_sheet = workbook["1-发票基本信息"]
        detail_sheet = workbook["2-发票明细信息"]
        basic_headers = self._header_map(basic_sheet)
        detail_headers = self._header_map(detail_sheet)

        self._clear_sheet_data(basic_sheet)
        self._clear_sheet_data(detail_sheet)

        for index, invoice in enumerate(invoices, start=DATA_START_ROW):
            if index > TEMPLATE_ROW:
                clone_row_style(basic_sheet, TEMPLATE_ROW, index, basic_sheet.max_column)
                clone_row_style(detail_sheet, TEMPLATE_ROW, index, detail_sheet.max_column)
            self._write_row(
                basic_sheet,
                index,
                basic_headers,
                {
                    "发票流水号": invoice.invoice_serial,
                    "发票类型": "普通发票",
                    "是否含税": "是",
                    "受票方自然人标识": "是" if invoice.is_natural_person else "否",
                    "购买方名称": invoice.invoice_title,
                    "购买方纳税人识别号": invoice.tax_id,
                    "购买方邮箱": invoice.email or None,
                    "备注": invoice.remark or None,
                },
            )
            self._write_row(
                detail_sheet,
                index,
                detail_headers,
                {
                    "发票流水号": invoice.invoice_serial,
                    "项目名称": "餐费",
                    "商品和服务税收编码": "3070401000000000000",
                    "金额": invoice.amount_text,
                    "税率": "0.01",
                },
            )

        ensure_parent_dir(backup_result.output_path)
        workbook.save(backup_result.output_path)
        return backup_result

    def _prepare_output(self, store: StoreConfig) -> BackupResult:
        output_path = store.output_xlsx_path
        ensure_parent_dir(output_path)
        backup_path: Optional[Path] = None
        if output_path.exists():
            timestamp = __import__("datetime").datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = self.backups_root.joinpath(store.store_key, f"{timestamp}.xlsx")
            ensure_parent_dir(backup_path)
            shutil.copy2(output_path, backup_path)
        return BackupResult(backup_path=backup_path, output_path=output_path)

    @staticmethod
    def _header_map(sheet: object) -> dict[str, int]:
        mapping: dict[str, int] = {}
        for column in range(1, sheet.max_column + 1):
            value = sheet.cell(HEADER_ROW, column).value
            if value:
                mapping[str(value).strip()] = column
        return mapping

    @staticmethod
    def _clear_sheet_data(sheet: object) -> None:
        for row in range(DATA_START_ROW, sheet.max_row + 1):
            for column in range(1, sheet.max_column + 1):
                sheet.cell(row, column).value = None

    @staticmethod
    def _write_row(sheet: object, row_index: int, header_map: dict[str, int], values: dict[str, Optional[str]]) -> None:
        for header, value in values.items():
            column = header_map.get(header)
            if not column:
                continue
            sheet.cell(row_index, column).value = value

